from __future__ import annotations

import argparse
import copy
import hashlib
import json
import random
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from difflib import SequenceMatcher
from typing import Any

import openpyxl


CATEGORY_LABELS = {
    1: "Hoteles / hosterias",
    2: "Apart / cabanas",
    3: "B&B / hospedajes",
    4: "CAT/DAT",
    5: "Hostels",
    6: "Campings",
}

SECTION_CATEGORY = {
    "hoteles y hosterias": 1,
    "apart cabanas": 2,
    "b b hospedaje residencial": 3,
    "cat dat casas y departamentos de alquiler turistico": 4,
    "albergues hostels hostales": 5,
    "campings": 6,
}


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def clean_string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_name(value: Any) -> str:
    text = normalize_text(value)
    for prefix in [
        "hotel ",
        "hosteria ",
        "hosterias ",
        "cabanas ",
        "cabana ",
        "apart ",
        "departamentos ",
        "departamento ",
        "camping ",
        "hostel ",
        "albergue ",
        "alojamiento ",
    ]:
        if text.startswith(prefix):
            text = text.removeprefix(prefix)
    return text.strip()


def similarity(left: str, right: str) -> float:
    score = SequenceMatcher(None, left, right).ratio()
    if len(left) >= 5 and len(right) >= 5 and (left in right or right in left):
        score = max(score, 0.92)
    return score


def parse_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else round(value)
    digits = re.sub(r"[^\d-]+", "", str(value))
    if not digits or digits == "-":
        return None
    try:
        return abs(int(digits))
    except ValueError:
        return None


def parse_category_numbers(value: Any, fallback: int) -> list[int]:
    numbers = [int(match) for match in re.findall(r"[1-6]", clean_string(value))]
    if not numbers:
        numbers = [fallback]
    return sorted(dict.fromkeys(numbers))


def normalize_phone_number(value: Any) -> tuple[str, list[str]]:
    raw = clean_string(value)
    if not raw:
        return "", []

    chunks = re.split(r"[/\n,;-]+", raw)
    normalized: list[str] = []
    for chunk in chunks:
        digits = re.sub(r"\D+", "", chunk)
        if not digits:
            continue
        if digits.startswith("00"):
            digits = digits[2:]
        if digits.startswith("0"):
            digits = digits[1:]
        if len(digits) == 7 and digits.startswith("4"):
            digits = f"2944{digits}"
        if not digits.startswith("54"):
            digits = f"54{digits}"
        normalized.append(f"+{digits}")

    normalized = list(dict.fromkeys(normalized))
    return (normalized[0] if normalized else ""), normalized


def phone_key(value: Any) -> str:
    digits = re.sub(r"\D+", "", "" if value is None else str(value))
    return digits[-8:] if len(digits) >= 8 else digits


def date_only(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


def date_document(value: str | None) -> dict[str, str] | None:
    if not value:
        return None
    return {"$date": f"{value}T00:00:00.000Z"}


def row_key(name: Any, habilitation: Any, social_reason: Any) -> tuple[str, str, str]:
    return (
        clean_string(parse_int(habilitation)),
        normalize_text(name),
        normalize_text(social_reason),
    )


def compact_key(name: Any, habilitation: Any, social_reason: Any) -> str:
    habilitation_text, name_text, social_text = row_key(name, habilitation, social_reason)
    return "|".join([habilitation_text, name_text, social_text])


def category_from_section(value: Any, default: int) -> int:
    normalized = normalize_text(value)
    return SECTION_CATEGORY.get(normalized, default)


def build_record(row: tuple[Any, ...], category: int, source_sheet: str, source_row: int) -> dict[str, Any] | None:
    name = clean_string(row[2] if len(row) > 2 else "")
    if not name:
        return None

    code = row[1] if len(row) > 1 else category
    phone, all_phones = normalize_phone_number(row[8] if len(row) > 8 else "")
    units = parse_int(row[10] if len(row) > 10 else None)
    places = parse_int(row[11] if len(row) > 11 else None)
    category_numbers = parse_category_numbers(code, category)
    accommodation_types = [CATEGORY_LABELS[number] for number in category_numbers]

    return {
        "source_sheet": source_sheet,
        "source_row": source_row,
        "category": clean_string(row[0] if len(row) > 0 else ""),
        "category_number": category,
        "category_numbers": category_numbers,
        "accommodation_type": CATEGORY_LABELS[category],
        "accommodation_types": accommodation_types,
        "accommodation_name": name,
        "establishment_name": name,
        "social_reason": clean_string(row[3] if len(row) > 3 else ""),
        "address": clean_string(row[4] if len(row) > 4 else ""),
        "habilitation_number": clean_string(row[5] if len(row) > 5 else ""),
        "parcel_number": clean_string(row[5] if len(row) > 5 else ""),
        "nomenclature": clean_string(row[6] if len(row) > 6 else ""),
        "neighborhood": clean_string(row[7] if len(row) > 7 else ""),
        "raw_phone": clean_string(row[8] if len(row) > 8 else ""),
        "phone": phone,
        "whatsapp": phone,
        "phones": all_phones,
        "email": clean_string(row[9] if len(row) > 9 else ""),
        "units": units,
        "places": places,
    }


def read_records(workbook: openpyxl.Workbook) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for sheet_name, default_category in [("ALOJAMIENTOS ", 1), ("CAMPINGS", 6)]:
        sheet = workbook[sheet_name]
        current_category = default_category
        for row_index in range(3, sheet.max_row + 1):
            row = tuple(sheet.cell(row_index, column).value for column in range(1, sheet.max_column + 1))
            if row[0]:
                current_category = category_from_section(row[0], default_category)
            if not clean_string(row[2] if len(row) > 2 else ""):
                continue
            category = 6 if sheet_name == "CAMPINGS" else current_category
            record = build_record(row, category, sheet_name.strip(), row_index)
            if record:
                records.append(record)
    return records


def read_temporary_leaves(workbook: openpyxl.Workbook) -> list[dict[str, Any]]:
    sheet = workbook["BAJAS TEMPORALES"]
    leaves: list[dict[str, Any]] = []
    ranges = [(3, 79, 4), (84, 94, 6)]
    for start, end, category_number in ranges:
        for row_index in range(start, end + 1):
            name = sheet.cell(row_index, 1).value
            if not clean_string(name):
                continue
            start_date = date_only(sheet.cell(row_index, 6).value)
            end_date = date_only(sheet.cell(row_index, 7).value)
            if not start_date or not end_date:
                continue
            leaves.append(
                {
                    "source_row": row_index,
                    "name": clean_string(name),
                    "social_reason": clean_string(sheet.cell(row_index, 2).value),
                    "address": clean_string(sheet.cell(row_index, 3).value),
                    "habilitation_number": clean_string(sheet.cell(row_index, 4).value),
                    "phone": clean_string(sheet.cell(row_index, 5).value),
                    "category_number": category_number,
                    "temporary_leave_start": start_date,
                    "temporary_leave_end": end_date,
                    "match_key": compact_key(name, sheet.cell(row_index, 4).value, sheet.cell(row_index, 2).value),
                }
            )
    return leaves


def apply_temporary_leaves(records: list[dict[str, Any]], leaves: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_habilitation: dict[str, list[dict[str, Any]]] = {}
    by_name: dict[str, list[dict[str, Any]]] = {}
    for record in records:
        habilitation = clean_string(parse_int(record.get("habilitation_number")))
        if habilitation:
            by_habilitation.setdefault(habilitation, []).append(record)
        by_name.setdefault(normalize_text(record["accommodation_name"]), []).append(record)

    unmatched = []
    for leave in leaves:
        candidates: list[dict[str, Any]] = []
        habilitation = clean_string(parse_int(leave["habilitation_number"]))
        if habilitation:
            candidates = by_habilitation.get(habilitation, [])
        if not candidates:
            candidates = by_name.get(normalize_text(leave["name"]), [])
        if leave["category_number"] == 6:
            candidates = [record for record in candidates if record["category_number"] == 6] or candidates
        else:
            candidates = [record for record in candidates if record["category_number"] != 6] or candidates
        if not candidates:
            unmatched.append(leave)
            continue
        for record in candidates:
            record["temporary_leave_start"] = leave["temporary_leave_start"]
            record["temporary_leave_end"] = leave["temporary_leave_end"]
            record["temporary_leave_source_row"] = leave["source_row"]
    return unmatched


def record_from_unmatched_leave(leave: dict[str, Any]) -> dict[str, Any]:
    phone, all_phones = normalize_phone_number(leave["phone"])
    category_number = leave["category_number"]
    return {
        "source_sheet": "BAJAS TEMPORALES",
        "source_row": leave["source_row"],
        "category": "BAJA TEMPORAL",
        "category_number": category_number,
        "category_numbers": [category_number],
        "accommodation_type": CATEGORY_LABELS[category_number],
        "accommodation_types": [CATEGORY_LABELS[category_number]],
        "accommodation_name": leave["name"],
        "establishment_name": leave["name"],
        "social_reason": leave["social_reason"],
        "address": leave["address"],
        "habilitation_number": leave["habilitation_number"],
        "parcel_number": leave["habilitation_number"],
        "nomenclature": "",
        "neighborhood": "",
        "raw_phone": leave["phone"],
        "phone": phone,
        "whatsapp": phone,
        "phones": all_phones,
        "email": "",
        "units": None,
        "places": None,
        "temporary_leave_start": leave["temporary_leave_start"],
        "temporary_leave_end": leave["temporary_leave_end"],
        "temporary_leave_source_row": leave["source_row"],
        "created_from_temporary_leave": True,
    }


def assign_ids(records: list[dict[str, Any]], seed: int) -> list[dict[str, Any]]:
    rng = random.Random(seed)
    used: set[str] = set()
    for record in records:
        while True:
            candidate = str(rng.randint(10_000_000, 99_999_999))
            if candidate not in used:
                used.add(candidate)
                break
        record["id"] = candidate
    return records


def build_user_documents(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    now = datetime.now(timezone.utc).isoformat()
    users = [
        {
            "_id": "meb-admin",
            "role": "admin",
            "display_name": "Admin MEB",
            "username": "admin",
            "password": "admin123",
            "seed_source": "mora_ultima_etapa_2026",
            "updated_at": now,
        }
    ]
    for record in records:
        document = {
            "_id": record["id"],
            "role": "establishment",
            "display_name": record["accommodation_name"],
            "establishment_name": record["establishment_name"],
            "accommodation_name": record["accommodation_name"],
            "social_reason": record["social_reason"],
            "parcel_number": record["parcel_number"],
            "address": record["address"],
            "phone": record["phone"],
            "whatsapp": record["whatsapp"],
            "raw_phone": record["raw_phone"],
            "phones": record["phones"],
            "email": record["email"],
            "units": record["units"],
            "places": record["places"],
            "accommodation_type": record["accommodation_type"],
            "accommodation_types": record["accommodation_types"],
            "category": record["category"],
            "category_number": record["category_number"],
            "category_numbers": record["category_numbers"],
            "habilitation_number": record["habilitation_number"],
            "nomenclature": record["nomenclature"],
            "neighborhood": record["neighborhood"],
            "temporary_leave_start": date_document(record.get("temporary_leave_start")),
            "temporary_leave_end": date_document(record.get("temporary_leave_end")),
            "temporary_leave_source_row": record.get("temporary_leave_source_row"),
            "source_sheet": record["source_sheet"],
            "source_row": record["source_row"],
            "seed_source": "mora_ultima_etapa_2026",
            "updated_at": now,
        }
        users.append({key: value for key, value in document.items() if value not in ("", [], None)})
    return users


def remap_historic_entries(users: list[dict[str, Any]], old_export_path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if not old_export_path.exists():
        return [], {"historic_export_found": False}

    old_data = json.loads(old_export_path.read_text(encoding="utf-8-sig"))
    old_users = [user for user in old_data["collections"]["users"] if user.get("role") == "establishment"]
    old_entries = [
        entry for entry in old_data["collections"].get("occupancy_entries", [])
        if "Importado desde" in (entry.get("notes") or "")
        and (
            "OCUPACION" in (entry.get("notes") or "").upper()
            or "OCUPACIÓN" in (entry.get("notes") or "").upper()
        )
    ]
    skipped_non_official_entries = len(old_data["collections"].get("occupancy_entries", [])) - len(old_entries)
    new_users = [user for user in users if user.get("role") == "establishment"]

    new_by_phone: dict[str, list[dict[str, Any]]] = defaultdict(list)
    new_names = []
    for user in new_users:
        for field in ["phone", "whatsapp", "raw_phone"]:
            key = phone_key(user.get(field))
            if key:
                new_by_phone[key].append(user)
        new_names.append((clean_name(user.get("accommodation_name") or user.get("establishment_name")), user))

    matches: dict[str, dict[str, Any]] = {}
    match_reasons: Counter[str] = Counter()
    unmatched_establishments: list[dict[str, Any]] = []
    for old_user in old_users:
        old_name = clean_name(old_user.get("accommodation_name") or old_user.get("establishment_name"))
        old_type = old_user.get("accommodation_type")
        candidates: list[dict[str, Any]] = []
        for field in ["phone", "whatsapp", "raw_phone"]:
            key = phone_key(old_user.get(field))
            if key:
                candidates.extend(new_by_phone.get(key, []))
        unique_candidates = {candidate["_id"]: candidate for candidate in candidates}
        candidates = list(unique_candidates.values())

        if len(candidates) > 1 and old_type:
            typed = [
                candidate for candidate in candidates
                if candidate.get("accommodation_type") == old_type or old_type in candidate.get("accommodation_types", [])
            ]
            if typed:
                candidates = typed

        if candidates:
            scored = sorted(
                [
                    (
                        similarity(old_name, clean_name(candidate.get("accommodation_name") or candidate.get("establishment_name"))),
                        candidate,
                    )
                    for candidate in candidates
                ],
                key=lambda item: item[0],
                reverse=True,
            )
            if len(scored) == 1 or scored[0][0] - scored[1][0] > 0.05:
                matches[old_user["_id"]] = scored[0][1]
                match_reasons["phone_or_type"] += 1
                continue

        scoped_names = [
            (name, user) for name, user in new_names
            if not old_type or user.get("accommodation_type") == old_type or old_type in user.get("accommodation_types", [])
        ] or new_names
        scored = sorted(
            [(similarity(old_name, name), user) for name, user in scoped_names],
            key=lambda item: item[0],
            reverse=True,
        )
        if scored and scored[0][0] >= 0.82 and (len(scored) == 1 or scored[0][0] - scored[1][0] > 0.03):
            matches[old_user["_id"]] = scored[0][1]
            match_reasons["name"] += 1
            continue
        unmatched_establishments.append(
            {
                "id": old_user["_id"],
                "name": old_user.get("accommodation_name") or old_user.get("establishment_name"),
                "entries": 0,
            }
        )

    entry_counts = Counter(entry.get("establishment_id") for entry in old_entries)
    for unmatched in unmatched_establishments:
        unmatched["entries"] = entry_counts.get(unmatched["id"], 0)

    aggregated: dict[tuple[str, str], dict[str, Any]] = {}
    unmatched_entries = 0
    for entry in old_entries:
        target = matches.get(entry.get("establishment_id"))
        if not target:
            unmatched_entries += 1
            continue
        week_start = entry["week_start"]["$date"] if isinstance(entry.get("week_start"), dict) else str(entry.get("week_start"))
        key = (target["_id"], week_start[:10])
        if key not in aggregated:
            next_entry = copy.deepcopy(entry)
            next_entry["_id"] = {"$oid": hashlib.md5(f"{key[0]}-{key[1]}".encode()).hexdigest()[:24]}
            next_entry["establishment_id"] = target["_id"]
            next_entry["establishment_name"] = target.get("establishment_name") or target.get("display_name")
            next_entry["notes"] = f"{entry.get('notes') or ''} Importado desde base historica y remapeado al padron Mora.".strip()
            aggregated[key] = next_entry
            continue
        aggregated[key]["occupied_places"] = (aggregated[key].get("occupied_places") or 0) + (entry.get("occupied_places") or 0)
        aggregated[key]["occupied_units"] = (aggregated[key].get("occupied_units") or 0) + (entry.get("occupied_units") or 0)

    entries = sorted(aggregated.values(), key=lambda entry: (entry["week_start"]["$date"], entry["establishment_name"]))
    return entries, {
        "historic_export_found": True,
        "old_establishments": len(old_users),
        "matched_establishments": len(matches),
        "unmatched_establishments": [
            item for item in unmatched_establishments if item["entries"] > 0
        ],
        "old_entries": len(old_entries),
        "skipped_non_official_entries": skipped_non_official_entries,
        "remapped_entries": len(entries),
        "unmatched_entries": unmatched_entries,
        "match_reasons": dict(match_reasons),
    }


def write_outputs(records: list[dict[str, Any]], users: list[dict[str, Any]], occupancy_entries: list[dict[str, Any]], unmatched: list[dict[str, Any]], occupancy_report: dict[str, Any]) -> None:
    seed_records = []
    for record in records:
        seed_record = {key: value for key, value in record.items() if key != "id"}
        seed_record["id"] = record["id"]
        seed_records.append(seed_record)
    Path("backend/app/establishments_seed.json").write_text(
        json.dumps(seed_records, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    occupancy_seed_records = []
    for entry in occupancy_entries:
        seed_entry = {key: value for key, value in entry.items() if key != "_id"}
        for field in ["week_start", "created_at", "updated_at"]:
            value = seed_entry.get(field)
            if isinstance(value, dict) and "$date" in value:
                seed_entry[field] = value["$date"]
        occupancy_seed_records.append(seed_entry)
    Path("backend/app/occupancy_seed.json").write_text(
        json.dumps(occupancy_seed_records, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    Path("exports").mkdir(exist_ok=True)
    Path("exports/turismo-mora-ultima-etapa-db.json").write_text(
        json.dumps(
            {
                "source": "LISTADO COMPLETO TURISMO MORA - ULTIMA ETAPA.xlsx",
                "collections": {
                    "users": users,
                    "occupancy_entries": occupancy_entries,
                },
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    Path("exports/turismo-mora-import-report.json").write_text(
        json.dumps(
            {
                "establishments": len(records),
                "temporary_leaves_applied": sum(1 for record in records if record.get("temporary_leave_start")),
                "temporary_leaves_unmatched": unmatched,
                "occupancy": occupancy_report,
                "by_category": {
                    CATEGORY_LABELS[number]: sum(1 for record in records if record["category_number"] == number)
                    for number in sorted(CATEGORY_LABELS)
                },
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", nargs="?", default=r"C:\Users\PC 1\Downloads\LISTADO COMPLETO TURISMO MORA - ÚLTIMA ETAPA.xlsx")
    parser.add_argument("--seed", type=int, default=20260807)
    parser.add_argument("--historic-export", default="exports/turismo-db-export-2026-06-08.json")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        matches = list(Path(r"C:\Users\PC 1\Downloads").glob("LISTADO COMPLETO TURISMO MORA*.xlsx"))
        if not matches:
            raise FileNotFoundError(workbook_path)
        workbook_path = matches[0]

    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    records = read_records(workbook)
    leaves = read_temporary_leaves(workbook)
    unmatched = apply_temporary_leaves(records, leaves)
    records.extend(record_from_unmatched_leave(leave) for leave in unmatched)
    records = assign_ids(records, args.seed)
    users = build_user_documents(records)
    occupancy_entries, occupancy_report = remap_historic_entries(users, Path(args.historic_export))
    write_outputs(records, users, occupancy_entries, [], occupancy_report)

    print(f"Establecimientos generados: {len(records)}")
    print(f"Usuarios Mongo a insertar: {len(users)}")
    print(f"Bajas temporales leidas con fecha: {len(leaves)}")
    print(f"Bajas temporales aplicadas: {sum(1 for record in records if record.get('temporary_leave_start'))}")
    print(f"Bajas temporales agregadas desde hoja de bajas: {len(unmatched)}")
    print(f"Cargas historicas remapeadas: {len(occupancy_entries)}")
    for number in sorted(CATEGORY_LABELS):
        print(f"- {CATEGORY_LABELS[number]}: {sum(1 for record in records if record['category_number'] == number)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
