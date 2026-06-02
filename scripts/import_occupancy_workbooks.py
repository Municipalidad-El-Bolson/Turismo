from __future__ import annotations

import argparse
import json
import math
import re
import sys
import unicodedata
from difflib import SequenceMatcher
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from urllib.error import HTTPError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

import openpyxl


API_URL = "http://localhost:8000"
ADMIN_ID = "meb-admin"

SOURCE_FILES = [
    r"C:\Users\PC 1\Downloads\2026.01 - OCUPACION DEL 02 AL 05 DE ENERO.xlsx",
    r"C:\Users\PC 1\Downloads\2026.01 - OCUPACION DEL 06 AL 11 DE ENERO.xlsx",
    r"C:\Users\PC 1\Downloads\2026.01 - OCUPACION DEL 12 AL 18 DE ENERO.xlsx",
    r"C:\Users\PC 1\Downloads\2026.01 - OCUPACION DEL 19 AL 25 DE ENERO.xlsx",
    r"C:\Users\PC 1\Downloads\2026.01 - OCUPACION DEL 26 AL 31 DE ENERO.xlsx",
    r"C:\Users\PC 1\Downloads\2026.02 - OCUPACION DEL 02 AL 07 DE febrero.xlsx",
    r"C:\Users\PC 1\Downloads\2026.02 - OCUPACION CARNAVAL DEL 13 AL 17 DE FEB.xlsx",
    r"C:\Users\PC 1\Downloads\2026.02 - OCUPACION DEL 19 AL 22 FIESTA DEL LUPULO.xlsx",
    r"C:\Users\PC 1\Downloads\2026.03 - OCUPACION FINDE XL 01 AL 04 MARZO.xlsx",
    r"C:\Users\PC 1\Downloads\2026.03 - OCUPACION DEL 12 AL 15 DE MARZO.xlsx",
    r"C:\Users\PC 1\Downloads\2026.03 - OCUPACION DEL 21 AL 23 DE MARZO.xlsx",
]

TYPE_BY_SHEET = {
    "HOTELHOSTERIASBYB": "Hoteles / hosterias",
    "CABANAS": "Apart / cabanas",
    "CABAÑAS": "Apart / cabanas",
    "HOSTELS": "Hostels",
    "CATDAT": "Apart / cabanas",
    "CAMPING": "Campings / dormis",
}

TYPE_BY_CODE = {
    1: "Hoteles / hosterias",
    2: "Apart / cabanas",
    3: "B&B / hospedajes",
    4: "Apart / cabanas",
    5: "Hostels",
    6: "Campings / dormis",
}


@dataclass
class ImportedRow:
    source_file: str
    sheet_name: str
    week_start: date
    establishment_name: str
    phone: str
    occupied_units: int
    occupied_places: int
    accommodation_type: str
    matched_id: str | None = None
    matched_name: str | None = None
    match_reason: str | None = None


def normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def normalize_name(value: object) -> str:
    text = normalize_text(value)
    replacements = {
        "hotel ": "",
        "hosteria ": "",
        "cabanas ": "",
        "cabana ": "",
        "apart ": "",
        "departamentos ": "",
        "departamento ": "",
        "camping ": "",
        "hostel ": "",
    }
    for old, new in replacements.items():
        if text.startswith(old):
            text = text.replace(old, new, 1)
    text = text.replace(" deptos", " departamentos").replace(" depto", " departamento")
    return text.strip()


def digits(value: object) -> str:
    if isinstance(value, int | float) and not isinstance(value, bool) and math.isfinite(value):
        value = str(int(value))
    return re.sub(r"\D+", "", "" if value is None else str(value))


def phone_key(value: object) -> str:
    number = digits(value)
    return number[-8:] if len(number) >= 8 else number


def phone_keys(value: object) -> list[str]:
    number = digits(value)
    keys = []
    if len(number) >= 8:
        keys.append(number[-8:])
    if len(number) >= 7:
        keys.append(number[-7:])
    elif number:
        keys.append(number)
    return list(dict.fromkeys(keys))


def parse_week_start(path: Path) -> date:
    match = re.search(r"(\d{4})\.(\d{2}).*?(?:DEL|CARNAVAL DEL|FINDE XL)\s+(\d{1,2})", path.name, re.I)
    if match:
        year, month, day = (int(part) for part in match.groups())
        return date(year, month, day)

    match = re.search(r"(\d{2})\.(\d{4})", path.name)
    if match:
        month, year = (int(part) for part in match.groups())
        return date(year, month, 1)

    raise ValueError(f"No se pudo leer la fecha del archivo: {path.name}")


def parse_monthly_file_date(path: Path) -> date:
    match = re.search(r"(\d{2})\.(\d{4})", path.name)
    if not match:
        return parse_week_start(path)
    month, year = (int(part) for part in match.groups())
    return date(year, month, 1)


def is_number(value: object) -> bool:
    return isinstance(value, int | float) and not isinstance(value, bool) and math.isfinite(value)


def average_int(values: list[float]) -> int:
    return int(round(sum(values) / len(values))) if values else 0


def sheet_type(sheet_name: str) -> str | None:
    return TYPE_BY_SHEET.get(normalize_text(sheet_name).upper())


def extract_day_pairs(header: tuple[object, ...]) -> list[tuple[int, int]]:
    pairs = []
    for index in range(4, len(header) - 1):
        left = normalize_text(header[index])
        right = normalize_text(header[index + 1])
        if left.startswith("u") and right.startswith("p"):
            pairs.append((index, index + 1))
    return pairs


def extract_rows(path: Path) -> list[ImportedRow]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    week_start = parse_week_start(path)
    rows: list[ImportedRow] = []

    if "DATOS DIARIOS" in workbook.sheetnames:
        daily_rows = extract_daily_rows(path, workbook)
        if daily_rows:
            return daily_rows

    if "DATOS X 15-MES" in workbook.sheetnames:
        monthly_rows = extract_monthly_rows(path, workbook)
        if monthly_rows:
            return monthly_rows

    for sheet in workbook.worksheets:
        accommodation_type = sheet_type(sheet.title)
        if not accommodation_type:
            continue

        all_rows = list(sheet.iter_rows(values_only=True))
        if not all_rows:
            continue

        pairs = extract_day_pairs(tuple(all_rows[0]))
        if not pairs:
            continue

        for row in all_rows[1:]:
            name = row[0] if len(row) > 0 else None
            if not isinstance(name, str) or not normalize_name(name):
                continue
            if normalize_text(name).startswith("ref"):
                continue
            if normalize_text(name) in {"semana", "total", "total unid", "total plazas"}:
                continue

            units = [float(row[u]) for u, _ in pairs if u < len(row) and is_number(row[u])]
            places = [float(row[p]) for _, p in pairs if p < len(row) and is_number(row[p])]
            if not units and not places:
                continue

            rows.append(
                ImportedRow(
                    source_file=path.name,
                    sheet_name=sheet.title,
                    week_start=week_start,
                    establishment_name=name.strip(),
                    phone=str(row[1]).strip() if len(row) > 1 and row[1] is not None else "",
                    occupied_units=average_int(units),
                    occupied_places=average_int(places),
                    accommodation_type=accommodation_type,
                )
            )
    return rows


def extract_monthly_rows(path: Path, workbook) -> list[ImportedRow]:
    sheet = workbook["DATOS X 15-MES"]
    week_start = parse_week_start(path)
    rows: list[ImportedRow] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[1] if len(row) > 1 else None
        if not isinstance(name, str) or not normalize_name(name):
            continue

        occupied_units = row[20] if len(row) > 20 else None
        occupied_places = row[21] if len(row) > 21 else None
        if not is_number(occupied_units) and not is_number(occupied_places):
            continue
        if (occupied_units or 0) == 0 and (occupied_places or 0) == 0:
            continue

        type_code = int(row[2]) if len(row) > 2 and is_number(row[2]) else 0
        rows.append(
            ImportedRow(
                source_file=path.name,
                sheet_name=sheet.title,
                week_start=week_start,
                establishment_name=name.strip(),
                phone="",
                occupied_units=int(round(float(occupied_units or 0))),
                occupied_places=int(round(float(occupied_places or 0))),
                accommodation_type=TYPE_BY_CODE.get(type_code, "Otros"),
            )
        )
    return rows


def extract_daily_rows(path: Path, workbook) -> list[ImportedRow]:
    sheet = workbook["DATOS DIARIOS"]
    month_start = parse_monthly_file_date(path)
    header = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    pairs: list[tuple[int, int, int]] = []
    for index in range(5, len(header) - 1):
        left = normalize_text(header[index])
        right = normalize_text(header[index + 1])
        match = re.match(r"u\s*(\d{1,2})$", left)
        if match and re.match(r"p\s*\d{1,2}$", right):
            pairs.append((index, index + 1, int(match.group(1))))

    rows: list[ImportedRow] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[1] if len(row) > 1 else None
        if not isinstance(name, str) or not normalize_name(name):
            continue

        type_code = int(row[2]) if len(row) > 2 and is_number(row[2]) else 0
        for unit_index, place_index, day in pairs:
            occupied_units = row[unit_index] if unit_index < len(row) else None
            occupied_places = row[place_index] if place_index < len(row) else None
            if not is_number(occupied_units) and not is_number(occupied_places):
                continue
            if (occupied_units or 0) == 0 and (occupied_places or 0) == 0:
                continue
            rows.append(
                ImportedRow(
                    source_file=path.name,
                    sheet_name=f"{sheet.title} dia {day:02d}",
                    week_start=date(month_start.year, month_start.month, day),
                    establishment_name=name.strip(),
                    phone="",
                    occupied_units=int(round(float(occupied_units or 0))),
                    occupied_places=int(round(float(occupied_places or 0))),
                    accommodation_type=TYPE_BY_CODE.get(type_code, "Otros"),
                )
            )
    return rows


def api_get(path: str) -> object:
    with urlopen(f"{API_URL}{path}", timeout=30) as response:
        return json.loads(response.read().decode("utf-8"))


def api_post(path: str, payload: dict) -> object:
    data = json.dumps(payload).encode("utf-8")
    request = Request(
        f"{API_URL}{path}",
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urlopen(request, timeout=30) as response:
        return json.loads(response.read().decode("utf-8"))


def load_establishments() -> list[dict]:
    query = urlencode({"userId": ADMIN_ID})
    return api_get(f"/establishments?{query}")  # type: ignore[return-value]


def build_indexes(establishments: list[dict]) -> tuple[dict[str, list[dict]], dict[str, list[dict]], dict[str, list[dict]]]:
    by_name: dict[str, list[dict]] = defaultdict(list)
    by_clean_name: dict[str, list[dict]] = defaultdict(list)
    by_phone: dict[str, list[dict]] = defaultdict(list)
    for establishment in establishments:
        names = [establishment.get("accommodation_name"), establishment.get("establishment_name")]
        for name in names:
            if name:
                add_unique(by_name[normalize_text(name)], establishment)
                add_unique(by_clean_name[normalize_name(name)], establishment)
        for key in phone_keys(establishment.get("phone") or establishment.get("whatsapp")):
            add_unique(by_phone[key], establishment)
    return by_name, by_clean_name, by_phone


def add_unique(items: list[dict], establishment: dict) -> None:
    if not any(item["id"] == establishment["id"] for item in items):
        items.append(establishment)


def match_rows(rows: list[ImportedRow], establishments: list[dict]) -> None:
    by_name, by_clean_name, by_phone = build_indexes(establishments)
    clean_candidates = []
    for establishment in establishments:
        clean_candidates.append(
            (
                normalize_name(establishment.get("accommodation_name") or establishment.get("establishment_name")),
                establishment,
            )
        )

    for row in rows:
        candidates = by_name.get(normalize_text(row.establishment_name), [])
        reason = "name"
        if len(candidates) != 1:
            candidates = by_clean_name.get(normalize_name(row.establishment_name), [])
            reason = "clean_name"
        if len(candidates) != 1:
            candidates = []
            for key in phone_keys(row.phone):
                for candidate in by_phone.get(key, []):
                    add_unique(candidates, candidate)
            reason = "phone"
        if len(candidates) > 1:
            typed = filter_by_type(candidates, row.accommodation_type)
            if len(typed) == 1:
                candidates = typed
                reason = "type"
        if len(candidates) > 1:
            candidates = best_name_match(normalize_name(row.establishment_name), candidates, 0.45)
            reason = "phone_name"
        if len(candidates) != 1:
            candidates = best_global_match(normalize_name(row.establishment_name), clean_candidates, 0.78, row.accommodation_type)
            reason = "fuzzy_type_name"
        if len(candidates) != 1:
            candidates = best_global_match(normalize_name(row.establishment_name), clean_candidates, 0.86, None)
            reason = "fuzzy_name"

        if len(candidates) == 1:
            row.matched_id = candidates[0]["id"]
            row.matched_name = candidates[0].get("accommodation_name") or candidates[0].get("establishment_name")
            row.match_reason = reason


def similarity(left: str, right: str) -> float:
    score = SequenceMatcher(None, left, right).ratio()
    if len(left) >= 5 and len(right) >= 5 and (left in right or right in left):
        score = max(score, 0.9)
    left_numbers = set(re.findall(r"\d+", left))
    right_numbers = set(re.findall(r"\d+", right))
    if left_numbers and left_numbers.intersection(right_numbers):
        score = max(score, 0.88)
    return score


def filter_by_type(candidates: list[dict], accommodation_type: str | None) -> list[dict]:
    if not accommodation_type:
        return candidates
    typed = [candidate for candidate in candidates if candidate.get("accommodation_type") == accommodation_type]
    return typed or candidates


def best_name_match(name: str, candidates: list[dict], threshold: float) -> list[dict]:
    scored = [
        (
            similarity(name, normalize_name(candidate.get("accommodation_name") or candidate.get("establishment_name"))),
            candidate,
        )
        for candidate in candidates
    ]
    scored.sort(key=lambda item: item[0], reverse=True)
    if not scored or scored[0][0] < threshold:
        return candidates
    if len(scored) > 1 and scored[0][0] - scored[1][0] < 0.08:
        return candidates
    return [scored[0][1]]


def best_global_match(
    name: str,
    candidates: list[tuple[str, dict]],
    threshold: float,
    accommodation_type: str | None,
) -> list[dict]:
    scoped = [
        (candidate_name, establishment)
        for candidate_name, establishment in candidates
        if not accommodation_type or establishment.get("accommodation_type") == accommodation_type
    ]
    scored = [(similarity(name, candidate_name), establishment) for candidate_name, establishment in scoped]
    scored.sort(key=lambda item: item[0], reverse=True)
    if not scored or scored[0][0] < threshold:
        return []
    if len(scored) > 1 and scored[0][0] - scored[1][0] < 0.08:
        return []
    return [scored[0][1]]


def import_rows(rows: list[ImportedRow]) -> Counter:
    counts: Counter = Counter()
    for row in aggregate_matched_rows(rows):
        if not row.matched_id:
            counts["skipped_unmatched"] += 1
            continue
        payload = {
            "week_start": row.week_start.isoformat(),
            "occupied_places": row.occupied_places,
            "occupied_units": row.occupied_units,
            "notes": f"Importado desde {row.source_file} ({row.sheet_name}).",
        }
        query = urlencode({"userId": ADMIN_ID})
        try:
            api_post(f"/establishments/{row.matched_id}/entries?{query}", payload)
            counts["imported"] += 1
        except HTTPError as exc:
            counts[f"http_{exc.code}"] += 1
    return counts


def aggregate_matched_rows(rows: list[ImportedRow]) -> list[ImportedRow]:
    aggregated: dict[tuple[str, date], ImportedRow] = {}
    unmatched = [row for row in rows if not row.matched_id]
    for row in rows:
        if not row.matched_id:
            continue
        key = (row.matched_id, row.week_start)
        existing = aggregated.get(key)
        if not existing:
            aggregated[key] = row
            continue
        existing.occupied_units += row.occupied_units
        existing.occupied_places += row.occupied_places
        existing.source_file = row.source_file
        existing.sheet_name = f"{existing.sheet_name}, {row.sheet_name}"
    return list(aggregated.values()) + unmatched


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--file", action="append", dest="files")
    args = parser.parse_args()

    paths = [Path(path) for path in (args.files or SOURCE_FILES)]
    missing = [path for path in paths if not path.exists()]
    if missing:
        print("Archivos faltantes:")
        for path in missing:
            print(f"- {path}")
        return 1

    rows = [row for path in paths for row in extract_rows(path)]
    establishments = load_establishments()
    match_rows(rows, establishments)

    matched = [row for row in rows if row.matched_id]
    unmatched = [row for row in rows if not row.matched_id]
    unique_imports = [row for row in aggregate_matched_rows(rows) if row.matched_id]
    by_file = Counter(row.source_file for row in rows)
    by_type = Counter(row.accommodation_type for row in rows)

    print(f"Filas extraidas: {len(rows)}")
    print(f"Filas con match: {len(matched)}")
    print(f"Filas sin match: {len(unmatched)}")
    print(f"Cargas unicas a importar: {len(unique_imports)}")
    print("Por archivo:")
    for file_name, count in by_file.items():
        print(f"- {file_name}: {count}")
    print("Por tipo:")
    for type_name, count in by_type.items():
        print(f"- {type_name}: {count}")

    if unmatched:
        print("Primeros sin match:")
        for row in unmatched[:25]:
            print(f"- {row.establishment_name} | {row.phone} | {row.source_file} | {row.sheet_name}")

    if args.dry_run:
        return 0

    counts = import_rows(rows)
    print("Resultado importacion:")
    for key, value in counts.items():
        print(f"- {key}: {value}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
