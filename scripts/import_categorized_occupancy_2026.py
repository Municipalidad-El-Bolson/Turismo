from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import openpyxl


DEFAULT_WORKBOOK = r"C:\Users\PC 1\Downloads\Ocupacion_alojamientos_categorizados_2026.xlsx"


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def parse_int(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, bool):
        return 0
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(round(value))
    digits = re.sub(r"[^\d-]+", "", str(value))
    if not digits or digits == "-":
        return 0
    return int(digits)


def parse_date_label(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    day, month, year = (int(part) for part in text.split("/"))
    return f"{year:04d}-{month:02d}-{day:02d}"


def build_establishment_index(records: list[dict[str, Any]]) -> dict[tuple[Any, ...], list[dict[str, Any]]]:
    index: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        name = normalize_text(record.get("accommodation_name"))
        keys = [
            (name, record.get("source_sheet"), str(record.get("category_number")), record.get("units"), record.get("places")),
            (name, record.get("source_sheet"), None, record.get("units"), record.get("places")),
            (name, None, None, record.get("units"), record.get("places")),
            (name, None, None, None, None),
        ]
        for key in keys:
            index[key].append(record)
    return index


def find_establishment(
    index: dict[tuple[Any, ...], list[dict[str, Any]]],
    name: Any,
    source_sheet: Any,
    category_code: Any,
    units: Any,
    places: Any,
) -> dict[str, Any]:
    normalized_name = normalize_text(name)
    category_text = "" if category_code is None else str(category_code).strip()
    category_number = category_text[0] if category_text and category_text[0].isdigit() else None
    keys = [
        (normalized_name, source_sheet, category_number, units, places),
        (normalized_name, source_sheet, None, units, places),
        (normalized_name, None, None, units, places),
        (normalized_name, None, None, None, None),
    ]
    for key in keys:
        matches = {record["id"]: record for record in index.get(key, [])}
        if len(matches) == 1:
            return next(iter(matches.values()))
        if len(matches) > 1:
            raise ValueError(f"Coincidencia ambigua para {name}: {[item['id'] for item in matches.values()]}")
    raise ValueError(f"No se encontro alojamiento para {name}")


def read_occupancy_entries(workbook_path: Path, establishments: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet = workbook["Ocupación 2026"]
    index = build_establishment_index(establishments)
    date_pairs = []
    for column in range(8, sheet.max_column + 1, 2):
        date_label = sheet.cell(4, column).value
        if not date_label:
            continue
        date_pairs.append((parse_date_label(date_label), column, column + 1))

    entries = []
    rows_with_data: set[str] = set()
    by_period: Counter[str] = Counter()
    by_category: Counter[str] = Counter()
    now = datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")

    for row in range(6, sheet.max_row + 1):
        name = sheet.cell(row, 4).value
        if not name:
            continue
        establishment = find_establishment(
            index,
            name,
            sheet.cell(row, 7).value,
            sheet.cell(row, 3).value,
            sheet.cell(row, 5).value,
            sheet.cell(row, 6).value,
        )
        for entry_date, unit_column, place_column in date_pairs:
            occupied_units = sheet.cell(row, unit_column).value
            occupied_places = sheet.cell(row, place_column).value
            if occupied_units is None and occupied_places is None:
                continue
            rows_with_data.add(establishment["id"])
            by_period[entry_date[:7]] += 1
            by_category[establishment.get("accommodation_type", "Sin categoria")] += 1
            entries.append(
                {
                    "establishment_id": establishment["id"],
                    "establishment_name": establishment["accommodation_name"],
                    "week_start": f"{entry_date}T00:00:00Z",
                    "occupied_places": parse_int(occupied_places),
                    "occupied_units": parse_int(occupied_units),
                    "notes": f"Importado desde {workbook_path.name} ({entry_date}).",
                    "created_at": now,
                    "updated_at": now,
                }
            )

    return entries, {
        "source": workbook_path.name,
        "entries": len(entries),
        "establishments_with_data": len(rows_with_data),
        "periods": dict(sorted(by_period.items())),
        "by_category": dict(sorted(by_category.items())),
    }


def ejson_entry(entry: dict[str, Any]) -> dict[str, Any]:
    key = f"{entry['establishment_id']}-{entry['week_start']}"
    return {
        "_id": {"$oid": hashlib.md5(key.encode()).hexdigest()[:24]},
        "establishment_id": entry["establishment_id"],
        "establishment_name": entry["establishment_name"],
        "week_start": {"$date": entry["week_start"]},
        "occupied_places": entry["occupied_places"],
        "occupied_units": entry["occupied_units"],
        "notes": entry["notes"],
        "created_at": {"$date": entry["created_at"]},
        "updated_at": {"$date": entry["updated_at"]},
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", nargs="?", default=DEFAULT_WORKBOOK)
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    establishments_path = Path("backend/app/establishments_seed.json")
    export_path = Path("exports/turismo-mora-ultima-etapa-db.json")
    report_path = Path("exports/turismo-mora-import-report.json")

    establishments = json.loads(establishments_path.read_text(encoding="utf-8"))
    entries, report = read_occupancy_entries(workbook_path, establishments)

    Path("backend/app/occupancy_seed.json").write_text(
        json.dumps(entries, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    export = json.loads(export_path.read_text(encoding="utf-8"))
    export["collections"]["occupancy_entries"] = [ejson_entry(entry) for entry in entries]
    export_path.write_text(json.dumps(export, ensure_ascii=False, indent=2), encoding="utf-8")

    existing_report = json.loads(report_path.read_text(encoding="utf-8"))
    existing_report["occupancy"] = report
    report_path.write_text(json.dumps(existing_report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Cargas generadas: {report['entries']}")
    print(f"Establecimientos con datos: {report['establishments_with_data']}")
    print(f"Periodos: {report['periods']}")
    print(f"Por categoria: {report['by_category']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
